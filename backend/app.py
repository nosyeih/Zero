from flask import Flask, render_template, session, redirect, url_for, request, flash
from authlib.integrations.flask_client import OAuth
from flask_sqlalchemy import SQLAlchemy
import os
import requests

from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "your_secret_key_here")
app.config['GOOGLE_CLIENT_ID'] = os.getenv("GOOGLE_CLIENT_ID")
app.config['GOOGLE_CLIENT_SECRET'] = os.getenv("GOOGLE_CLIENT_SECRET")

# Database Configuration
# Handle Render's postgres:// -> postgresql://
database_url = os.getenv("DATABASE_URL", "sqlite:///users.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# User Model
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=True)
    picture = db.Column(db.String(200), nullable=True)
    webapp_url = db.Column(db.String(500), nullable=True) # Stores the user's Sheet URL

    def __repr__(self):
        return f'<User {self.email}>'

# Initialize DB
with app.app_context():
    db.create_all()

oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=app.config['GOOGLE_CLIENT_ID'],
    client_secret=app.config['GOOGLE_CLIENT_SECRET'],
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'},
)

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login')
def login():
    redirect_uri = url_for('auth', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/auth/callback')
def auth():
    token = google.authorize_access_token()
    user_info = google.parse_id_token(token, nonce=None)
    
    email = user_info.get('email')
    name = user_info.get('name')
    picture = user_info.get('picture')

    # Find or Create User
    user = User.query.filter_by(email=email).first()
    if not user:
        user = User(email=email, name=name, picture=picture)
        db.session.add(user)
        db.session.commit()
    else:
        # Update info if changed
        user.name = name
        user.picture = picture
        db.session.commit()

    session['user_id'] = user.id
    return redirect(url_for('dashboard'))

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('index'))

@app.route('/setup')
def setup():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    return render_template('setup.html', user=user)

@app.route('/save_setup', methods=['POST'])
def save_setup():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    webapp_url = request.form.get('webapp_url')
    
    if webapp_url:
        user.webapp_url = webapp_url
        db.session.commit()
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('dashboard'))
    
    flash('Por favor ingresa una URL válida.', 'error')
    return redirect(url_for('setup'))

import json
from collections import defaultdict
from datetime import datetime

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    return render_template('hub.html', user=user)

@app.route('/operations')
def operations():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    
    if not user.webapp_url:
        return redirect(url_for('setup'))

    # Fetch data from Google Sheet
    transactions = []
    sheet_status = 'error'
    
    try:
        response = requests.get(user.webapp_url, allow_redirects=True, timeout=5)
        if response.status_code == 200:
            data_resp = response.json()
            if isinstance(data_resp, dict) and 'data' in data_resp:
                transactions = data_resp['data']
            elif isinstance(data_resp, list):
                transactions = data_resp
            else:
                transactions = []
            sheet_status = 'success'
    except Exception as e:
        flash(f"Error de conexión con la hoja: {str(e)}", 'error')
        sheet_status = 'error'

    # Process Data for Dashboard
    totals = {
        'PEN': {'income': 0, 'expense': 0, 'balance': 0},
        'USD': {'income': 0, 'expense': 0, 'balance': 0}
    }
    
    # Structure for charts: {'Month': {'Ingreso': 0, 'Egreso': 0}}
    chart_data_pen = defaultdict(lambda: {'Ingreso': 0, 'Egreso': 0})
    chart_data_usd = defaultdict(lambda: {'Ingreso': 0, 'Egreso': 0})

    def parse_float(value):
        """Robustly parse float from mixed string formats"""
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s: return 0.0
        
        # Handle European format: 1.234,56
        if '.' in s and ',' in s:
            if s.rfind(',') > s.rfind('.'):
                # Formato 1.234,56 -> quitar puntos, reemplazar coma por punto
                s = s.replace('.', '').replace(',', '.')
            else:
                # Formato 1,234.56 -> quitar comas
                s = s.replace(',', '')
        elif ',' in s:
            # Caso 12,50 o 1,200
            # Si hay más de una coma, asumo miles: 1,234,567 -> quitar
            # Si hay una sola coma y está al final (decimal): 12,50 -> punto
            if s.count(',') == 1:
                # Heurística: si son 3 dígitos después de coma, podría ser miles (1,000)
                # Pero en finanzas (dinero) usualmente son 2 decimales. 
                # Asumiremos coma = decimal si parece decimal
                parts = s.split(',')
                if len(parts[1]) == 2: 
                    s = s.replace(',', '.')
                else: 
                     s = s.replace(',', '') # asumo miles
            else:
                s = s.replace(',', '')
        
        try:
            return float(s)
        except:
            return 0.0

    def parse_date_key(value):
        """Parse date to YYYY-MM"""
        s = str(value).strip()
        if not s: return 'Unknown'
        
        # Clean up potential time components or extra whitespace
        # Example: "2025-12-31T00:00:00" -> "2025-12-31" inside try blocks or split
        s_date_part = s.split(' ')[0].split('T')[0]

        # List of formats to try
        formats = [
            '%Y-%m-%d',  # 2025-12-31
            '%d/%m/%Y',  # 15/01/2026 or 15/1/2026
            '%d-%m-%Y',  # 15-01-2026
            '%Y/%m/%d',  # 2025/12/31
            '%d/%m/%y',  # 15/01/26
        ]

        for fmt in formats:
            try:
                return datetime.strptime(s_date_part, fmt).strftime('%Y-%m')
            except ValueError:
                continue
        
        # If specific simplistic slicing worked before, keep it as fallback/context
        # But for '16/1/2016' slicing first 10 chars is fine: '16/1/2016' (9 chars)
        
        return 'Unknown'
    
    # Debug: Print first transaction to logs to verify structure
    if transactions:
        print(f"DEBUG - First Transaction: {transactions[0]}", flush=True)

    for t in transactions:
        try:
            monto = parse_float(t.get('MONTO', 0))
            
            # Robust Currency Parsing
            raw_moneda = str(t.get('MONEDA', '')).upper().strip()
            if 'USD' in raw_moneda:
                moneda = 'USD'
            elif 'PEN' in raw_moneda or 'SOL' in raw_moneda: 
                moneda = 'PEN'
            else:
                moneda = 'PEN' # Default fallback
            
            # Parse Month Key for Charts
            month_key = parse_date_key(t.get('FECHA_INGRESO', ''))
            
            # Parse ISO Date for Frontend Sorting/Display
            # We try to get a clean YYYY-MM-DD for the table
            raw_date = str(t.get('FECHA_INGRESO', '')).strip()
            parsed_iso_date = raw_date # Fallback
            
            # reuse the logic from parse_date_key but return full date
            try:
                # Cleanup
                s_date = raw_date.split(' ')[0].split('T')[0]
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y']:
                    try:
                        dt_obj = datetime.strptime(s_date, fmt)
                        parsed_iso_date = dt_obj.strftime('%Y-%m-%d')
                        break
                    except:
                        continue
            except:
                pass

            # Inject parsed values into the dictionary for the Frontend
            t['_parsed_monto'] = monto
            t['_parsed_moneda'] = moneda
            t['_parsed_date'] = parsed_iso_date
            t['_parsed_month'] = month_key # YYYY-MM

            if month_key == 'Unknown':
                 print(f"DEBUG - Unparseable Date: {t.get('FECHA_INGRESO')}", flush=True)

            val = monto
            if val >= 0:
                totals[moneda]['income'] += val
                if month_key != 'Unknown':
                    if moneda == 'PEN':
                        chart_data_pen[month_key]['Ingreso'] += val
                    else:
                        chart_data_usd[month_key]['Ingreso'] += val
            else:
                totals[moneda]['expense'] += abs(val)
                if month_key != 'Unknown':
                    if moneda == 'PEN':
                        chart_data_pen[month_key]['Egreso'] += abs(val)
                    else:
                        chart_data_usd[month_key]['Egreso'] += abs(val)

        except Exception as e:
            print(f"Error extracting row: {e}")
            continue

    totals['PEN']['balance'] = totals['PEN']['income'] - totals['PEN']['expense']
    totals['USD']['balance'] = totals['USD']['income'] - totals['USD']['expense']
    
    # Sort chart data for PEN
    sorted_months_pen = sorted(chart_data_pen.keys())
    labels_pen = sorted_months_pen
    income_data_pen = [chart_data_pen[m]['Ingreso'] for m in sorted_months_pen]
    expense_data_pen = [chart_data_pen[m]['Egreso'] for m in sorted_months_pen]

    # Sort chart data for USD
    sorted_months_usd = sorted(chart_data_usd.keys())
    labels_usd = sorted_months_usd
    income_data_usd = [chart_data_usd[m]['Ingreso'] for m in sorted_months_usd]
    expense_data_usd = [chart_data_usd[m]['Egreso'] for m in sorted_months_usd]

    return render_template('operations.html', user=user, transactions=transactions, sheet_status=sheet_status,
                           totals=totals,
                           chart_pen={'labels': labels_pen, 'income': income_data_pen, 'expense': expense_data_pen},
                           chart_usd={'labels': labels_usd, 'income': income_data_usd, 'expense': expense_data_usd})

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    if 'user_id' not in session:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
             return {"status": "error", "message": "Unauthorized"}, 401
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    
    if not user.webapp_url:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
             return {"status": "error", "message": "Setup required"}, 400
        return redirect(url_for('setup'))

    # Support both Form Data and JSON
    if request.is_json:
        req_data = request.get_json()
        fecha_ingreso = req_data.get('fecha_ingreso')
        empresa = req_data.get('empresa')
        concepto_pago = req_data.get('concepto_pago')
        moneda = req_data.get('moneda')
        monto = float(req_data.get('monto'))
        nro_operacion = req_data.get('nro_operacion')
        detalle = req_data.get('detalle')
    else:
        fecha_ingreso = request.form.get('fecha_ingreso')
        empresa = request.form.get('empresa')
        concepto_pago = request.form.get('concepto_pago')
        moneda = request.form.get('moneda')
        monto = float(request.form.get('monto'))
        nro_operacion = request.form.get('nro_operacion')
        detalle = request.form.get('detalle')

    data = {
        'fecha_ingreso': fecha_ingreso,
        'empresa': empresa,
        'concepto_pago': concepto_pago,
        'moneda': moneda,
        'monto': monto,
        'nro_operacion': nro_operacion,
        'detalle': detalle
    }

    try:
        response = requests.post(user.webapp_url, json=data, allow_redirects=True)
        
        if response.status_code == 200 or response.status_code == 302:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return {"status": "success", "message": "Operación registrada con éxito."}
            flash("Operación registrada con éxito.", 'success')
        else:
            msg = f"Error al registrar: {response.text}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return {"status": "error", "message": msg}, 500
            flash(msg, 'error')
             
    except Exception as e:
        msg = f"Error de conexión: {str(e)}"
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
            return {"status": "error", "message": msg}, 500
        flash(msg, 'error')
        
    return redirect(url_for('dashboard'))

import pandas as pd
import math
import openpyxl
from openpyxl.utils.cell import coordinate_from_string

@app.route('/profit', methods=['GET', 'POST'])
def profit_calculator():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    
    analysis = None
    error = None

    if request.method == 'POST':
        if 'excel_file' not in request.files:
            error = "No se subió ningún archivo."
            return render_template('profit.html', user=user, analysis=None, error=error)
            
        file = request.files['excel_file']
        if file.filename == '':
            error = "No seleccionaste archivo."
            return render_template('profit.html', user=user, analysis=None, error=error)
            
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            error = "Formato inválido. Sube un archivo Excel."
            return render_template('profit.html', user=user, analysis=None, error=error)

        try:
            # Get mapping config
            sli = float(request.form.get('sli', 0))
            col_name = request.form.get('col_name', '').strip()
            col_qty = request.form.get('col_qty', '').strip()
            col_unit_cost = request.form.get('col_unit_cost', '').strip()
            col_total_cost = request.form.get('col_total_cost', '').strip()

            # Function to split 'A2' into ('A', 2)
            def parse_coord(coord_str):
                try:
                    col, row = coordinate_from_string(coord_str)
                    return col, row
                except:
                    return None, None

            n_col, n_row = parse_coord(col_name)
            q_col, q_row = parse_coord(col_qty)
            u_col, u_row = parse_coord(col_unit_cost)
            t_col, t_row = parse_coord(col_total_cost)

            if not all([n_col, q_col, u_col, t_col]):
                 error = "Formato de coordenada inválida. Use formato como A2, B5."
                 return render_template('profit.html', user=user, analysis=None, error=error)
            
            # Load workbook with openpyxl (data only so we get formula results)
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active # Take the first/active sheet

            total_fob = 0
            processed_items = []
            
            # We assume the user gave us the starting row for the first product.
            start_row = n_row
            max_row = ws.max_row
            
            for r in range(start_row, max_row + 1):
                try:
                    name_val = ws[f"{n_col}{r}"].value
                    qty_val = ws[f"{q_col}{r}"].value
                    unit_val = ws[f"{u_col}{r}"].value
                    total_val = ws[f"{t_col}{r}"].value
                    
                    # Stop parsing if all critical fields are empty
                    if name_val is None and qty_val is None and unit_val is None:
                        break
                        
                    # Clean NaNs/None
                    if qty_val is None or unit_val is None or total_val is None:
                        continue
                        
                    qty = float(qty_val)
                    unit_fob = float(unit_val)
                    row_total = float(total_val)
                    name = str(name_val).strip() if name_val else f"Item {r}"

                    if qty > 0 and unit_fob > 0:
                        total_fob += row_total
                        processed_items.append({
                            'name': name,
                            'qty': qty,
                            'unit_fob': unit_fob,
                            'total_fob': row_total
                        })
                except Exception as e:
                    print(f"Skipping row {r} due to parse error: {e}")
                    continue

            # Calculate Factor
            factor = 0
            if total_fob > 0:
                factor = sli / total_fob
            
            # Calculate Landed Cost
            final_items = []
            for item in processed_items:
                unit_landed = item['unit_fob'] * (1 + factor)
                item['unit_landed'] = unit_landed
                final_items.append(item)

            analysis = {
                'total_fob': total_fob,
                'sli': sli,
                'factor': factor,
                'products': final_items
            }

            return render_template('profit.html', user=user, analysis=analysis, error=None)

        except Exception as e:
            import traceback
            traceback.print_exc()
            error = f"Error procesando el Excel: {str(e)}"
            return render_template('profit.html', user=user, analysis=None, error=error)

    # GET request
    return render_template('profit.html', user=user, analysis=None, error=None)

@app.route('/utilities')
def utilities():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user = db.session.get(User, session['user_id'])
    return render_template('utilities.html', user=user)

@app.route('/help')
def help_center():
    return render_template('help.html')

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')

@app.route('/terms')
def terms():
    return render_template('terms.html')




if __name__ == '__main__':
    app.run(debug=True)
